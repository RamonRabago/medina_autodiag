"""
Router para exportar reportes a Excel.
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime, date

from app.database import get_db
from app.models.venta import Venta
from app.models.detalle_venta import DetalleVenta
from app.models.cliente import Cliente
from app.models.orden_trabajo import OrdenTrabajo
from app.models.orden_compra import OrdenCompra, EstadoOrdenCompra
from app.models.pago_orden_compra import PagoOrdenCompra
from app.models.proveedor import Proveedor
from app.models.movimiento_inventario import MovimientoInventario, TipoMovimiento
from app.models.pago import Pago
from app.models.vehiculo import Vehiculo
from app.models.venta import Venta
from app.models.servicio import Servicio
from app.models.repuesto import Repuesto
from app.models.ubicacion import Ubicacion
from app.models.gasto_operativo import GastoOperativo
from app.services.gastos_service import query_gastos, CATEGORIAS_VALIDAS
from app.services.devoluciones_service import query_devoluciones
from app.models.caja_turno import CajaTurno
from app.models.usuario import Usuario
from app.models.cuenta_pagar_manual import CuentaPagarManual, PagoCuentaPagarManual
from app.models.comision_devengada import ComisionDevengada
from app.models.auditoria import Auditoria
from app.models.asistencia import Asistencia
from sqlalchemy import or_
from app.utils.roles import require_roles

router = APIRouter(
    prefix="/exportaciones",
    tags=["Exportaciones"]
)


def _encabezado(ws, titulos: list):
    """Estilo de encabezado."""
    for col, titulo in enumerate(titulos, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")


@router.get("/clientes")
def exportar_clientes(
    buscar: str | None = Query(None, description="Filtrar por nombre, teléfono, email, RFC"),
    limit: int = Query(5000, ge=1, le=10000),
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("ADMIN", "EMPLEADO", "TECNICO", "CAJA"))
):
    """Exporta el listado completo de clientes a Excel."""
    query = db.query(Cliente)
    if buscar and buscar.strip():
        term = f"%{buscar.strip()}%"
        filters = [
            Cliente.nombre.like(term),
            Cliente.telefono.like(term),
            Cliente.email.like(term),
            Cliente.direccion.like(term),
        ]
        if hasattr(Cliente, "rfc"):
            filters.append(Cliente.rfc.like(term))
        query = query.filter(or_(*filters))
    clientes = query.order_by(Cliente.nombre.asc()).limit(limit).all()

    # Contar ventas y vehículos por cliente
    ventas_count = {}
    vehiculos_count = {}
    for c in clientes:
        ventas_count[c.id_cliente] = db.query(Venta).filter(Venta.id_cliente == c.id_cliente).count()
        vehiculos_count[c.id_cliente] = db.query(Vehiculo).filter(Vehiculo.id_cliente == c.id_cliente).count()

    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    _encabezado(ws, ["ID", "Nombre", "Teléfono", "Email", "Dirección", "RFC", "Ventas", "Vehículos", "Fecha alta"])

    for row, c in enumerate(clientes, 2):
        ws.cell(row=row, column=1, value=c.id_cliente)
        ws.cell(row=row, column=2, value=c.nombre or "")
        ws.cell(row=row, column=3, value=c.telefono or "")
        ws.cell(row=row, column=4, value=c.email or "")
        ws.cell(row=row, column=5, value=(c.direccion or "")[:200])
        ws.cell(row=row, column=6, value=getattr(c, "rfc", None) or "")
        ws.cell(row=row, column=7, value=ventas_count.get(c.id_cliente, 0))
        ws.cell(row=row, column=8, value=vehiculos_count.get(c.id_cliente, 0))
        ws.cell(row=row, column=9, value=c.creado_en.strftime("%Y-%m-%d %H:%M") if c.creado_en else "")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"clientes_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"},
    )


@router.get("/ventas")
def exportar_ventas(
    fecha_desde: str | None = Query(None),
    fecha_hasta: str | None = Query(None),
    limit: int = Query(1000, ge=1, le=5000),
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("ADMIN", "EMPLEADO", "CAJA"))
):
    query = db.query(Venta)
    if fecha_desde:
        query = query.filter(func.date(Venta.fecha) >= fecha_desde)
    if fecha_hasta:
        query = query.filter(func.date(Venta.fecha) <= fecha_hasta)
    ventas = query.order_by(Venta.fecha.desc()).limit(limit).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"
    _encabezado(ws, ["ID", "Fecha", "Cliente", "Total", "Saldo pendiente", "Estado"])

    for row, v in enumerate(ventas, 2):
        cliente = db.query(Cliente).filter(Cliente.id_cliente == v.id_cliente).first() if v.id_cliente else None
        total_pagado = db.query(func.coalesce(func.sum(Pago.monto), 0)).filter(Pago.id_venta == v.id_venta).scalar()
        saldo = max(0, float(v.total) - float(total_pagado or 0))
        estado = v.estado.value if hasattr(v.estado, "value") else str(v.estado)
        ws.cell(row=row, column=1, value=v.id_venta)
        ws.cell(row=row, column=2, value=v.fecha.strftime("%Y-%m-%d %H:%M") if v.fecha else "")
        ws.cell(row=row, column=3, value=cliente.nombre if cliente else "-")
        ws.cell(row=row, column=4, value=float(v.total))
        ws.cell(row=row, column=5, value=round(saldo, 2))
        ws.cell(row=row, column=6, value=estado)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"ventas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"},
    )


@router.get("/productos-vendidos")
def exportar_productos_vendidos(
    fecha_desde: str | None = Query(None),
    fecha_hasta: str | None = Query(None),
    limit: int = Query(1000, ge=1, le=5000),
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("ADMIN", "EMPLEADO", "CAJA"))
):
    subq = (
        db.query(
            DetalleVenta.id_item,
            DetalleVenta.descripcion,
            func.sum(DetalleVenta.cantidad).label("cantidad"),
            func.sum(DetalleVenta.subtotal).label("monto"),
        )
        .join(Venta, Venta.id_venta == DetalleVenta.id_venta)
        .filter(DetalleVenta.tipo == "PRODUCTO", Venta.estado != "CANCELADA")
    )
    if fecha_desde:
        subq = subq.filter(func.date(Venta.fecha) >= fecha_desde)
    if fecha_hasta:
        subq = subq.filter(func.date(Venta.fecha) <= fecha_hasta)
    rows = (
        subq.group_by(DetalleVenta.id_item, DetalleVenta.descripcion)
        .order_by(func.sum(DetalleVenta.cantidad).desc())
        .limit(limit)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos más vendidos"
    _encabezado(ws, ["Producto", "Cantidad", "Monto"])

    for row, r in enumerate(rows, 2):
        ws.cell(row=row, column=1, value=r.descripcion or f"ID {r.id_item}")
        ws.cell(row=row, column=2, value=float(r.cantidad or 0))
        ws.cell(row=row, column=3, value=float(r.monto or 0))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"productos_vendidos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"},
    )


@router.get("/inventario")
def exportar_inventario(
    buscar: str | None = Query(None, description="Buscar por código, nombre o marca"),
    id_categoria: int | None = Query(None, description="Filtrar por categoría"),
    stock_bajo: bool = Query(False, description="Solo repuestos con stock bajo"),
    activo: bool | None = Query(None, description="Filtrar por activo/inactivo"),
    incluir_eliminados: bool = Query(False, description="Incluir productos eliminados"),
    limit: int = Query(5000, ge=1, le=10000),
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("ADMIN", "EMPLEADO", "TECNICO", "CAJA"))
):
    """Exporta el inventario de repuestos a Excel."""
    query = db.query(Repuesto).options(
        joinedload(Repuesto.categoria),
        joinedload(Repuesto.proveedor),
        joinedload(Repuesto.ubicacion_obj).joinedload(Ubicacion.bodega),
        joinedload(Repuesto.estante).joinedload(Estante.ubicacion).joinedload(Ubicacion.bodega),
        joinedload(Repuesto.nivel),
        joinedload(Repuesto.fila),
    )
    if not incluir_eliminados or getattr(current_user, "rol", None) != "ADMIN":
        query = query.filter(Repuesto.eliminado == False)
    if buscar and buscar.strip():
        term = f"%{buscar.strip()}%"
        query = query.filter(
            or_(
                Repuesto.codigo.like(term),
                Repuesto.nombre.like(term),
                Repuesto.marca.like(term),
            )
        )
    if id_categoria is not None:
        query = query.filter(Repuesto.id_categoria == id_categoria)
    if stock_bajo:
        query = query.filter(Repuesto.stock_actual <= Repuesto.stock_minimo)
    if activo is not None:
        query = query.filter(Repuesto.activo == activo)
    repuestos = query.order_by(Repuesto.codigo.asc()).limit(limit).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    _encabezado(ws, [
        "ID", "Código", "Nombre", "Categoría", "Proveedor", "Bodega", "Ubicación", "Stock", "Stock mín.", "Stock máx.",
        "P. compra", "P. venta", "Marca", "Unidad", "Estado", "Eliminado"
    ])

    for row, r in enumerate(repuestos, 2):
        cat_nombre = r.categoria.nombre if r.categoria else ""
        prov_nombre = r.proveedor.nombre if r.proveedor else ""
        bodega_nom = getattr(r, "bodega_nombre", "") or ""
        ubi_nom = getattr(r, "ubicacion_nombre", "") or r.ubicacion or ""
        estado = "Eliminado" if getattr(r, "eliminado", False) else ("Activo" if r.activo else "Inactivo")
        ws.cell(row=row, column=1, value=r.id_repuesto)
        ws.cell(row=row, column=2, value=r.codigo or "")
        ws.cell(row=row, column=3, value=r.nombre or "")
        ws.cell(row=row, column=4, value=cat_nombre)
        ws.cell(row=row, column=5, value=prov_nombre)
        ws.cell(row=row, column=6, value=bodega_nom)
        ws.cell(row=row, column=7, value=ubi_nom)
        ws.cell(row=row, column=8, value=r.stock_actual or 0)
        ws.cell(row=row, column=9, value=r.stock_minimo or 0)
        ws.cell(row=row, column=10, value=r.stock_maximo or 0)
        ws.cell(row=row, column=11, value=float(r.precio_compra or 0))
        ws.cell(row=row, column=12, value=float(r.precio_venta or 0))
        ws.cell(row=row, column=13, value=r.marca or "")
        ws.cell(row=row, column=14, value=r.unidad_medida or "PZA")
        ws.cell(row=row, column=15, value=estado)
        ws.cell(row=row, column=16, value="Sí" if getattr(r, "eliminado", False) else "")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"inventario_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"},
    )


@router.get("/servicios")
def exportar_servicios(
    buscar: str | None = Query(None, description="Buscar en código o nombre"),
    categoria: int | None = Query(None, description="Filtrar por id de categoría"),
    activo: bool | None = Query(None, description="Filtrar por activo/inactivo"),
    limit: int = Query(5000, ge=1, le=10000),
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("ADMIN", "EMPLEADO", "TECNICO", "CAJA"))
):
    """Exporta el catálogo de servicios a Excel."""
    query = db.query(Servicio)
    if buscar and buscar.strip():
        term = f"%{buscar.strip()}%"
        query = query.filter(
            (Servicio.codigo.like(term)) | (Servicio.nombre.like(term))
        )
    if categoria is not None:
        query = query.filter(Servicio.id_categoria == categoria)
    if activo is not None:
        query = query.filter(Servicio.activo == activo)
    servicios = query.options(joinedload(Servicio.categoria)).order_by(Servicio.codigo.asc()).limit(limit).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Servicios"
    _encabezado(ws, ["ID", "Código", "Nombre", "Categoría", "Descripción", "Precio", "Tiempo (min)", "Requiere repuestos", "Estado"])

    for row, s in enumerate(servicios, 2):
        cat_nombre = s.categoria.nombre if s.categoria else ""
        ws.cell(row=row, column=1, value=s.id)
        ws.cell(row=row, column=2, value=s.codigo or "")
        ws.cell(row=row, column=3, value=s.nombre or "")
        ws.cell(row=row, column=4, value=cat_nombre)
        ws.cell(row=row, column=5, value=(s.descripcion or "")[:200])
        ws.cell(row=row, column=6, value=float(s.precio_base or 0))
        ws.cell(row=row, column=7, value=s.tiempo_estimado_minutos or 0)
        ws.cell(row=row, column=8, value="Sí" if s.requiere_repuestos else "No")
        ws.cell(row=row, column=9, value="Activo" if s.activo else "Inactivo")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"servicios_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"},
    )


@router.get("/vehiculos")
def exportar_vehiculos(
    buscar: str | None = Query(None, description="Buscar en marca, modelo, VIN"),
    id_cliente: int | None = Query(None, description="Filtrar por cliente"),
    limit: int = Query(5000, ge=1, le=10000),
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("ADMIN", "EMPLEADO", "TECNICO", "CAJA"))
):
    """Exporta el listado completo de vehículos a Excel."""
    query = db.query(Vehiculo)
    if id_cliente:
        query = query.filter(Vehiculo.id_cliente == id_cliente)
    if buscar and buscar.strip():
        term = f"%{buscar.strip()}%"
        filters = [
            Vehiculo.marca.like(term),
            Vehiculo.modelo.like(term),
            Vehiculo.vin.like(term),
            Vehiculo.motor.like(term),
        ]
        if hasattr(Vehiculo, "color"):
            filters.append(Vehiculo.color.like(term))
        query = query.filter(or_(*filters))
    vehiculos = query.order_by(Vehiculo.id_vehiculo.desc()).limit(limit).all()

    def _color_display(v):
        c = getattr(v, "color", None)
        m = getattr(v, "motor", None)
        if c:
            return c
        if m and not str(m).replace(".", "").replace(",", "").isdigit():
            return m
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Vehículos"
    _encabezado(ws, ["ID", "Cliente", "Marca", "Modelo", "Año", "Color", "VIN", "Motor", "Fecha alta"])

    for row, v in enumerate(vehiculos, 2):
        cliente = db.query(Cliente).filter(Cliente.id_cliente == v.id_cliente).first() if v.id_cliente else None
        color_val = _color_display(v)
        ws.cell(row=row, column=1, value=v.id_vehiculo)
        ws.cell(row=row, column=2, value=cliente.nombre if cliente else "")
        ws.cell(row=row, column=3, value=v.marca or "")
        ws.cell(row=row, column=4, value=v.modelo or "")
        ws.cell(row=row, column=5, value=v.anio or "")
        ws.cell(row=row, column=6, value=color_val or "")
        ws.cell(row=row, column=7, value=v.vin or "")
        ws.cell(row=row, column=8, value=v.motor or "")
        ws.cell(row=row, column=9, value=v.creado_en.strftime("%Y-%m-%d %H:%M") if v.creado_en else "")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"vehiculos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"},
    )


@router.get("/clientes-frecuentes")
def exportar_clientes_frecuentes(
    fecha_desde: str | None = Query(None),
    fecha_hasta: str | None = Query(None),
    limit: int = Query(1000, ge=1, le=5000),
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("ADMIN", "EMPLEADO", "CAJA"))
):
    subq = (
        db.query(
            Venta.id_cliente,
            func.count(Venta.id_venta).label("ventas"),
            func.sum(Venta.total).label("total"),
        )
        .filter(Venta.estado != "CANCELADA", Venta.id_cliente.isnot(None))
    )
    if fecha_desde:
        subq = subq.filter(func.date(Venta.fecha) >= fecha_desde)
    if fecha_hasta:
        subq = subq.filter(func.date(Venta.fecha) <= fecha_hasta)
    rows = (
        subq.group_by(Venta.id_cliente)
        .order_by(func.count(Venta.id_venta).desc())
        .limit(limit)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes frecuentes"
    _encabezado(ws, ["Cliente", "Ventas", "Total"])

    for row, r in enumerate(rows, 2):
        c = db.query(Cliente).filter(Cliente.id_cliente == r.id_cliente).first()
        ws.cell(row=row, column=1, value=c.nombre if c else f"Cliente #{r.id_cliente}")
        ws.cell(row=row, column=2, value=r.ventas)
        ws.cell(row=row, column=3, value=float(r.total or 0))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"clientes_frecuentes_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"},
    )


@router.get("/cuentas-por-cobrar")
def exportar_cuentas_por_cobrar(
    fecha_desde: str | None = Query(None),
    fecha_hasta: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("ADMIN", "EMPLEADO", "CAJA"))
):
    query = db.query(Venta).filter(Venta.estado == "PENDIENTE")
    if fecha_desde:
        query = query.filter(func.date(Venta.fecha) >= fecha_desde)
    if fecha_hasta:
        query = query.filter(func.date(Venta.fecha) <= fecha_hasta)
    ventas = query.order_by(Venta.fecha.desc()).all()

    items = []
    for v in ventas:
        total_pagado = db.query(func.coalesce(func.sum(Pago.monto), 0)).filter(Pago.id_venta == v.id_venta).scalar()
        saldo = max(0, float(v.total) - float(total_pagado or 0))
        if saldo <= 0:
            continue
        cliente = db.query(Cliente).filter(Cliente.id_cliente == v.id_cliente).first() if v.id_cliente else None
        items.append({
            "id_venta": v.id_venta,
            "nombre_cliente": cliente.nombre if cliente else "-",
            "total": float(v.total),
            "saldo_pendiente": round(saldo, 2),
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "Cuentas por cobrar"
    _encabezado(ws, ["ID", "Cliente", "Total", "Saldo pendiente"])

    for row, it in enumerate(items, 2):
        ws.cell(row=row, column=1, value=it["id_venta"])
        ws.cell(row=row, column=2, value=it["nombre_cliente"])
        ws.cell(row=row, column=3, value=it["total"])
        ws.cell(row=row, column=4, value=it["saldo_pendiente"])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"cuentas_por_cobrar_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"},
    )


@router.get("/utilidad")
def exportar_utilidad(
    fecha_desde: str | None = Query(None),
    fecha_hasta: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("ADMIN", "CAJA")),
):
    """Exporta reporte de utilidad (ingresos - costo) a Excel."""
    query = db.query(Venta).filter(Venta.estado != "CANCELADA")
    if fecha_desde:
        query = query.filter(func.date(Venta.fecha) >= fecha_desde)
    if fecha_hasta:
        query = query.filter(func.date(Venta.fecha) <= fecha_hasta)
    ventas = query.order_by(Venta.fecha.asc()).all()

    from app.models.cancelacion_producto import CancelacionProducto

    total_ingresos = 0.0
    total_costo = 0.0
    filas = []

    for v in ventas:
        ingresos = float(v.total)
        costo = 0.0
        if v.id_orden:
            orden = db.query(OrdenTrabajo).filter(OrdenTrabajo.id == v.id_orden).first()
            if orden and not getattr(orden, "cliente_proporciono_refacciones", False):
                res = db.query(func.coalesce(func.sum(MovimientoInventario.costo_total), 0)).filter(
                    MovimientoInventario.tipo_movimiento == TipoMovimiento.SALIDA,
                    MovimientoInventario.referencia == orden.numero_orden,
                ).scalar()
                costo = float(res or 0)
        else:
            res = db.query(func.coalesce(func.sum(MovimientoInventario.costo_total), 0)).filter(
                MovimientoInventario.id_venta == v.id_venta,
                MovimientoInventario.tipo_movimiento == TipoMovimiento.SALIDA,
            ).scalar()
            costo = float(res or 0)
        utilidad = ingresos - costo
        total_ingresos += ingresos
        total_costo += costo
        filas.append((v.id_venta, v.fecha.strftime("%Y-%m-%d") if v.fecha else "", ingresos, costo, utilidad))

    perdidas_mer = 0.0
    query_cancel = db.query(Venta).filter(Venta.estado == "CANCELADA")
    if fecha_desde:
        query_cancel = query_cancel.filter(func.date(Venta.fecha) >= fecha_desde)
    if fecha_hasta:
        query_cancel = query_cancel.filter(func.date(Venta.fecha) <= fecha_hasta)
    ids_canceladas = [v.id_venta for v in query_cancel.all()]
    if ids_canceladas:
        res_mer = db.query(func.coalesce(func.sum(CancelacionProducto.costo_total_mer), 0)).filter(
            CancelacionProducto.id_venta.in_(ids_canceladas)
        ).scalar()
        perdidas_mer = float(res_mer or 0)

    utilidad_bruta = total_ingresos - total_costo - perdidas_mer

    total_gastos = 0.0
    q_gastos = db.query(GastoOperativo)
    if fecha_desde:
        q_gastos = q_gastos.filter(GastoOperativo.fecha >= fecha_desde)
    if fecha_hasta:
        q_gastos = q_gastos.filter(GastoOperativo.fecha <= fecha_hasta)
    res_gastos = q_gastos.with_entities(func.coalesce(func.sum(GastoOperativo.monto), 0)).scalar()
    total_gastos = float(res_gastos or 0)

    utilidad_neta = utilidad_bruta - total_gastos

    wb = Workbook()
    ws = wb.active
    ws.title = "Utilidad"
    _encabezado(ws, ["ID Venta", "Fecha", "Ingresos", "Costo", "Utilidad"])
    for row, (id_v, fch, ing, cos, util) in enumerate(filas, 2):
        ws.cell(row=row, column=1, value=id_v)
        ws.cell(row=row, column=2, value=fch)
        ws.cell(row=row, column=3, value=round(ing, 2))
        ws.cell(row=row, column=4, value=round(cos, 2))
        ws.cell(row=row, column=5, value=round(util, 2))
    row_total = len(filas) + 2
    ws.cell(row=row_total, column=1, value="SUBTOTAL VENTAS")
    ws.cell(row=row_total, column=3, value=round(total_ingresos, 2))
    ws.cell(row=row_total, column=4, value=round(total_costo, 2))
    ws.cell(row=row_total, column=5, value=round(total_ingresos - total_costo, 2))
    row_total += 1
    ws.cell(row=row_total, column=1, value="Pérdidas por merma (cancelaciones)")
    ws.cell(row=row_total, column=5, value=round(-perdidas_mer, 2))
    row_total += 1
    ws.cell(row=row_total, column=1, value="UTILIDAD BRUTA")
    ws.cell(row=row_total, column=5, value=round(utilidad_bruta, 2))
    row_total += 1
    ws.cell(row=row_total, column=1, value="Gastos operativos")
    ws.cell(row=row_total, column=5, value=round(-total_gastos, 2))
    row_total += 1
    ws.cell(row=row_total, column=1, value="UTILIDAD NETA")
    ws.cell(row=row_total, column=5, value=round(utilidad_neta, 2))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"utilidad_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"},
    )


@router.get("/comisiones")
def exportar_comisiones(
    fecha_desde: str | None = Query(None),
    fecha_hasta: str | None = Query(None),
    id_usuario: int | None = Query(None),
    limit: int = Query(5000, ge=1, le=50000),
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("ADMIN", "CAJA", "EMPLEADO", "TECNICO")),
):
    """Exporta reporte de comisiones devengadas a Excel. EMPLEADO/TECNICO: solo sus propias comisiones."""
    q = db.query(
        ComisionDevengada.id_usuario,
        ComisionDevengada.id_venta,
        ComisionDevengada.tipo_base,
        ComisionDevengada.base_monto,
        ComisionDevengada.porcentaje,
        ComisionDevengada.monto_comision,
        ComisionDevengada.fecha_venta,
    )
    if current_user.rol in ("EMPLEADO", "TECNICO"):
        q = q.filter(ComisionDevengada.id_usuario == current_user.id_usuario)
    elif id_usuario:
        q = q.filter(ComisionDevengada.id_usuario == id_usuario)
    if fecha_desde:
        q = q.filter(ComisionDevengada.fecha_venta >= fecha_desde)
    if fecha_hasta:
        q = q.filter(ComisionDevengada.fecha_venta <= fecha_hasta)
    rows = q.order_by(ComisionDevengada.fecha_venta.desc(), ComisionDevengada.id_venta).limit(limit).all()

    usuarios_cache = {}
    def _nombre(id_u):
        if id_u not in usuarios_cache:
            u = db.query(Usuario).filter(Usuario.id_usuario == id_u).first()
            usuarios_cache[id_u] = u.nombre if u else f"Usuario #{id_u}"
        return usuarios_cache[id_u]

    wb = Workbook()
    ws = wb.active
    ws.title = "Comisiones"
    _encabezado(ws, ["Empleado", "ID Venta", "Tipo base", "Base ($)", " % ", "Comisión ($)", "Fecha"])
    for row, r in enumerate(rows, 2):
        ws.cell(row=row, column=1, value=_nombre(r.id_usuario))
        ws.cell(row=row, column=2, value=r.id_venta)
        ws.cell(row=row, column=3, value=str(r.tipo_base))
        ws.cell(row=row, column=4, value=float(r.base_monto or 0))
        ws.cell(row=row, column=5, value=float(r.porcentaje or 0))
        ws.cell(row=row, column=6, value=float(r.monto_comision or 0))
        ws.cell(row=row, column=7, value=r.fecha_venta.strftime("%Y-%m-%d") if r.fecha_venta else "")

    total = sum(float(r.monto_comision or 0) for r in rows)
    row_total = len(rows) + 2
    ws.cell(row=row_total, column=1, value="TOTAL")
    ws.cell(row=row_total, column=6, value=round(total, 2))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"comisiones_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"},
    )


def _calcular_total_a_pagar_oc(oc):
    """Total a pagar de una orden de compra."""
    from decimal import Decimal
    total = Decimal("0")
    for d in oc.detalles:
        if d.cantidad_recibida <= 0:
            continue
        precio = d.precio_unitario_real if d.precio_unitario_real is not None else d.precio_unitario_estimado
        total += Decimal(str(d.cantidad_recibida)) * Decimal(str(precio))
    return float(total)


@router.get("/cuentas-por-pagar")
def exportar_cuentas_por_pagar(
    id_proveedor: int | None = Query(None),
    fecha_desde: str | None = Query(None, description="Recepción desde (YYYY-MM-DD)"),
    fecha_hasta: str | None = Query(None, description="Recepción hasta (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("ADMIN", "CAJA")),
):
    """Exporta cuentas por pagar (órdenes de compra con saldo pendiente) a Excel."""
    query = db.query(OrdenCompra).filter(
        OrdenCompra.estado.in_([
            EstadoOrdenCompra.RECIBIDA,
            EstadoOrdenCompra.RECIBIDA_PARCIAL,
        ]),
    )
    if id_proveedor:
        query = query.filter(OrdenCompra.id_proveedor == id_proveedor)
    ordenes = query.order_by(OrdenCompra.fecha.desc()).all()

    filas = []
    total_saldo = 0.0
    for oc in ordenes:
        total_a_pagar = _calcular_total_a_pagar_oc(oc)
        if total_a_pagar <= 0:
            continue
        pagos = db.query(PagoOrdenCompra).filter(
            PagoOrdenCompra.id_orden_compra == oc.id_orden_compra
        ).all()
        total_pagado = sum(float(p.monto) for p in pagos)
        saldo = max(0, total_a_pagar - total_pagado)
        if saldo <= 0:
            continue
        fch_rec = oc.fecha_recepcion.date() if oc.fecha_recepcion else None
        if fecha_desde or fecha_hasta:
            if not fch_rec:
                continue
            if fecha_desde and fch_rec < datetime.strptime(fecha_desde[:10], "%Y-%m-%d").date():
                continue
            if fecha_hasta and fch_rec > datetime.strptime(fecha_hasta[:10], "%Y-%m-%d").date():
                continue
        prov = db.query(Proveedor).filter(Proveedor.id_proveedor == oc.id_proveedor).first()
        hoy = date.today()
        dias = (hoy - fch_rec).days if fch_rec else None
        if dias is not None:
            antiguedad_rango = "0-30" if dias <= 30 else ("31-60" if dias <= 60 else "61+")
        else:
            antiguedad_rango = "-"
        filas.append((
            oc.numero,
            prov.nombre if prov else "",
            round(total_a_pagar, 2),
            round(total_pagado, 2),
            round(saldo, 2),
            oc.fecha_recepcion.strftime("%Y-%m-%d") if oc.fecha_recepcion else "",
            dias if dias is not None else "",
            antiguedad_rango,
        ))
        total_saldo += saldo

    wb = Workbook()
    ws = wb.active
    ws.title = "Cuentas por pagar"
    _encabezado(ws, ["Orden", "Proveedor", "Total a pagar", "Pagado", "Saldo pendiente", "Fecha recepción", "Días desde recepción", "Antigüedad"])
    for row, (num, prov, tot, pag, sal, fch, dias, ant) in enumerate(filas, 2):
        ws.cell(row=row, column=1, value=num)
        ws.cell(row=row, column=2, value=prov)
        ws.cell(row=row, column=3, value=tot)
        ws.cell(row=row, column=4, value=pag)
        ws.cell(row=row, column=5, value=sal)
        ws.cell(row=row, column=6, value=fch)
        ws.cell(row=row, column=7, value=dias)
        ws.cell(row=row, column=8, value=ant)
    row_total = len(filas) + 2
    ws.cell(row=row_total, column=1, value="TOTAL")
    ws.cell(row=row_total, column=5, value=round(total_saldo, 2))
    ws.cell(row=row_total + 1, column=1, value="Cuentas")
    ws.cell(row=row_total + 1, column=5, value=len(filas))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"cuentas_por_pagar_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"},
    )


@router.get("/cuentas-pagar-manuales")
def exportar_cuentas_pagar_manuales(
    id_proveedor: int | None = Query(None),
    fecha_desde: str | None = Query(None),
    fecha_hasta: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("ADMIN", "CAJA")),
):
    """Exporta cuentas por pagar manuales a Excel."""
    query = db.query(CuentaPagarManual).filter(CuentaPagarManual.cancelada == False)
    if id_proveedor:
        query = query.filter(CuentaPagarManual.id_proveedor == id_proveedor)
    cuentas = query.order_by(CuentaPagarManual.fecha_registro.desc()).all()
    filas = []
    total_saldo = 0.0
    hoy = date.today()
    for c in cuentas:
        total_pagado = sum(float(p.monto) for p in c.pagos)
        saldo = max(0, float(c.monto_total) - total_pagado)
        if saldo <= 0:
            continue
        if fecha_desde or fecha_hasta:
            f = c.fecha_registro
            if not f:
                continue
            if fecha_desde and f < datetime.strptime(fecha_desde[:10], "%Y-%m-%d").date():
                continue
            if fecha_hasta and f > datetime.strptime(fecha_hasta[:10], "%Y-%m-%d").date():
                continue
        nombre = c.proveedor.nombre if c.id_proveedor and c.proveedor else (c.acreedor_nombre or "")
        fch_ref = c.fecha_vencimiento or c.fecha_registro
        dias = (hoy - fch_ref).days if fch_ref else None
        ant = "0-30" if dias is not None and dias <= 30 else ("31-60" if dias is not None and dias <= 60 else ("61+" if dias is not None else "-"))
        filas.append((
            c.concepto,
            nombre,
            getattr(c, "referencia_factura", None) or "",
            round(float(c.monto_total), 2),
            round(total_pagado, 2),
            round(saldo, 2),
            c.fecha_registro.strftime("%Y-%m-%d") if c.fecha_registro else "",
            c.fecha_vencimiento.strftime("%Y-%m-%d") if c.fecha_vencimiento else "",
            dias if dias is not None else "",
            ant,
        ))
        total_saldo += saldo
    wb = Workbook()
    ws = wb.active
    ws.title = "Cuentas por pagar manuales"
    _encabezado(ws, ["Concepto", "Proveedor/Acreedor", "Ref. factura", "Total", "Pagado", "Saldo pendiente", "Fecha registro", "Vencimiento", "Días", "Antigüedad"])
    for row, tup in enumerate(filas, 2):
        for col, val in enumerate(tup, 1):
            ws.cell(row=row, column=col, value=val)
    r = len(filas) + 2
    ws.cell(row=r, column=1, value="TOTAL")
    ws.cell(row=r, column=6, value=round(total_saldo, 2))
    ws.cell(row=r + 1, column=1, value="Cuentas")
    ws.cell(row=r + 1, column=6, value=len(filas))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"cuentas_pagar_manuales_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"},
    )


@router.get("/auditoria")
def exportar_auditoria(
    fecha_desde: str | None = Query(None, description="Fecha desde (YYYY-MM-DD)"),
    fecha_hasta: str | None = Query(None, description="Fecha hasta (YYYY-MM-DD)"),
    modulo: str | None = Query(None, description="Filtrar por módulo"),
    id_usuario: int | None = Query(None, description="Filtrar por usuario"),
    limit: int = Query(2000, ge=1, le=10000),
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("ADMIN", "CAJA")),
):
    """Exporta el registro de auditoría (acciones de usuarios) a Excel."""
    from datetime import timedelta
    query = db.query(Auditoria).options(joinedload(Auditoria.usuario))
    if fecha_desde:
        try:
            f = datetime.strptime(fecha_desde[:10], "%Y-%m-%d")
            query = query.filter(Auditoria.fecha >= f)
        except (ValueError, TypeError):
            pass
    if fecha_hasta:
        try:
            f = datetime.strptime(fecha_hasta[:10], "%Y-%m-%d")
            f_end = f + timedelta(days=1)
            query = query.filter(Auditoria.fecha < f_end)
        except (ValueError, TypeError):
            pass
    if modulo:
        query = query.filter(Auditoria.modulo.ilike(f"%{modulo}%"))
    if id_usuario:
        query = query.filter(Auditoria.id_usuario == id_usuario)
    registros = query.order_by(Auditoria.fecha.desc()).limit(limit).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Auditoría"
    _encabezado(ws, ["Fecha", "Usuario", "Módulo", "Acción", "ID Referencia", "Descripción"])
    for row, r in enumerate(registros, 2):
        u = r.usuario if hasattr(r, "usuario") and r.usuario else None
        usu_nombre = (u.nombre or u.email or "") if u else ""
        ws.cell(row=row, column=1, value=r.fecha.strftime("%Y-%m-%d %H:%M:%S") if r.fecha else "")
        ws.cell(row=row, column=2, value=usu_nombre)
        ws.cell(row=row, column=3, value=r.modulo or "")
        ws.cell(row=row, column=4, value=r.accion or "")
        ws.cell(row=row, column=5, value=r.id_referencia)
        ws.cell(row=row, column=6, value=(r.descripcion or "")[:500])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"auditoria_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"},
    )


@router.get("/ajustes-inventario")
def exportar_ajustes_inventario(
    fecha_desde: str | None = Query(None, description="Fecha desde (YYYY-MM-DD)"),
    fecha_hasta: str | None = Query(None, description="Fecha hasta (YYYY-MM-DD)"),
    id_usuario: int | None = Query(None, description="Filtrar por usuario"),
    limit: int = Query(5000, ge=1, le=20000),
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("ADMIN", "CAJA")),
):
    """Exporta el reporte de ajustes de inventario (conteo físico/auditoría) a Excel."""
    query = db.query(MovimientoInventario).filter(
        MovimientoInventario.tipo_movimiento.in_(
            [TipoMovimiento.AJUSTE_POSITIVO, TipoMovimiento.AJUSTE_NEGATIVO]
        )
    ).options(
        joinedload(MovimientoInventario.repuesto),
        joinedload(MovimientoInventario.usuario),
    )
    if fecha_desde:
        query = query.filter(func.date(MovimientoInventario.fecha_movimiento) >= fecha_desde)
    if fecha_hasta:
        query = query.filter(func.date(MovimientoInventario.fecha_movimiento) <= fecha_hasta)
    if id_usuario:
        query = query.filter(MovimientoInventario.id_usuario == id_usuario)

    movimientos = query.order_by(MovimientoInventario.fecha_movimiento.desc()).limit(limit).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Ajustes inventario"
    _encabezado(ws, [
        "Fecha", "Repuesto", "Código", "Tipo", "Cantidad", "Stock ant.", "Stock nuevo",
        "Costo total", "Usuario", "Referencia / Motivo"
    ])

    for row, m in enumerate(movimientos, 2):
        rep_nombre = m.repuesto.nombre if m.repuesto else ""
        rep_codigo = m.repuesto.codigo if m.repuesto else ""
        usu_nombre = m.usuario.nombre if m.usuario else ""
        costo = float(m.costo_total) if m.costo_total is not None else 0
        ref_motivo = (m.referencia or "") + (" – " + m.motivo if m.motivo else "")
        ws.cell(row=row, column=1, value=m.fecha_movimiento.strftime("%Y-%m-%d %H:%M") if m.fecha_movimiento else "")
        ws.cell(row=row, column=2, value=rep_nombre)
        ws.cell(row=row, column=3, value=rep_codigo)
        ws.cell(row=row, column=4, value=m.tipo_movimiento.value if m.tipo_movimiento else "")
        ws.cell(row=row, column=5, value=m.cantidad)
        ws.cell(row=row, column=6, value=m.stock_anterior)
        ws.cell(row=row, column=7, value=m.stock_nuevo)
        ws.cell(row=row, column=8, value=round(costo, 2))
        ws.cell(row=row, column=9, value=usu_nombre)
        ws.cell(row=row, column=10, value=ref_motivo[:500])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"ajustes_inventario_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"},
    )


@router.get("/devoluciones")
def exportar_devoluciones(
    fecha_desde: str | None = Query(None, description="Fecha desde (YYYY-MM-DD)"),
    fecha_hasta: str | None = Query(None, description="Fecha hasta (YYYY-MM-DD)"),
    buscar: str | None = Query(None, description="Buscar en repuesto, referencia o motivo"),
    tipo_motivo: str | None = Query(
        None,
        description="Filtrar: venta (devolución por venta) u orden (cancelación de orden)",
    ),
    limit: int = Query(5000, ge=1, le=20000),
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("ADMIN", "CAJA", "TECNICO")),
):
    """Exporta el listado de devoluciones al inventario a Excel."""
    query = query_devoluciones(
        db,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        buscar=buscar,
        tipo_motivo=tipo_motivo,
    )
    movimientos = (
        query.options(
            joinedload(MovimientoInventario.repuesto),
            joinedload(MovimientoInventario.usuario),
        )
        .order_by(MovimientoInventario.fecha_movimiento.desc())
        .limit(limit)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Devoluciones"
    _encabezado(ws, ["Fecha", "Repuesto", "Código", "Cantidad", "Motivo", "Referencia"])

    for row, m in enumerate(movimientos, 2):
        rep_nombre = m.repuesto.nombre if m.repuesto else ""
        rep_codigo = m.repuesto.codigo if m.repuesto else ""
        fecha = m.fecha_movimiento.strftime("%Y-%m-%d %H:%M") if m.fecha_movimiento else ""
        ref = f"Venta #{m.id_venta}" if m.id_venta else (m.referencia or "")
        ws.cell(row=row, column=1, value=fecha)
        ws.cell(row=row, column=2, value=rep_nombre)
        ws.cell(row=row, column=3, value=rep_codigo)
        ws.cell(row=row, column=4, value=m.cantidad)
        ws.cell(row=row, column=5, value=(m.motivo or "")[:500])
        ws.cell(row=row, column=6, value=ref)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"devoluciones_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"},
    )


@router.get("/gastos")
def exportar_gastos(
    fecha_desde: str | None = Query(None, description="Fecha desde (YYYY-MM-DD)"),
    fecha_hasta: str | None = Query(None, description="Fecha hasta (YYYY-MM-DD)"),
    categoria: str | None = Query(None, description="RENTA, SERVICIOS, MATERIAL, NOMINA, OTROS, DEVOLUCION_VENTA"),
    buscar: str | None = Query(None, description="Buscar en concepto"),
    limit: int = Query(5000, ge=1, le=20000),
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("ADMIN", "CAJA")),
):
    """Exporta el listado de gastos operativos a Excel."""
    if categoria and categoria.strip().upper() not in CATEGORIAS_VALIDAS:
        raise HTTPException(
            status_code=400,
            detail=f"Categoría inválida: '{categoria}'. Use: {', '.join(CATEGORIAS_VALIDAS)}",
        )
    query = query_gastos(
        db,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        categoria=categoria,
        buscar=buscar,
    )
    gastos = query.order_by(GastoOperativo.fecha.desc()).limit(limit).all()

    CAT_LABELS = {"RENTA": "Renta", "SERVICIOS": "Servicios", "MATERIAL": "Material", "NOMINA": "Nómina", "DEVOLUCION_VENTA": "Devolución venta", "OTROS": "Otros"}

    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos operativos"
    _encabezado(ws, ["Fecha", "Concepto", "Categoría", "Monto", "Observaciones"])

    for row, g in enumerate(gastos, 2):
        ws.cell(row=row, column=1, value=g.fecha.strftime("%Y-%m-%d") if g.fecha else "")
        ws.cell(row=row, column=2, value=(g.concepto or "")[:200])
        ws.cell(row=row, column=3, value=CAT_LABELS.get(g.categoria, g.categoria))
        ws.cell(row=row, column=4, value=round(float(g.monto or 0), 2))
        ws.cell(row=row, column=5, value=(g.observaciones or "")[:500])

    row_total = len(gastos) + 2
    total_monto = sum(float(g.monto or 0) for g in gastos)
    ws.cell(row=row_total, column=1, value="TOTAL")
    ws.cell(row=row_total, column=4, value=round(total_monto, 2))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"gastos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"},
    )


@router.get("/asistencia")
def exportar_asistencia(
    fecha_desde: str = Query(..., description="Fecha desde (YYYY-MM-DD)"),
    fecha_hasta: str = Query(..., description="Fecha hasta (YYYY-MM-DD)"),
    id_usuario: int | None = Query(None, description="Filtrar por empleado"),
    limit: int = Query(5000, ge=1, le=20000),
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("ADMIN", "CAJA", "TECNICO", "EMPLEADO")),
):
    """Exporta registros de asistencia a Excel por rango de fechas. Requiere fecha_desde y fecha_hasta."""
    rol = getattr(current_user.rol, "value", None) or str(current_user.rol)
    if rol in ("TECNICO", "EMPLEADO"):
        id_usuario = current_user.id_usuario
    fd = datetime.strptime(fecha_desde[:10], "%Y-%m-%d").date()
    fh = datetime.strptime(fecha_hasta[:10], "%Y-%m-%d").date()
    query = db.query(Asistencia).filter(
        Asistencia.fecha >= fd,
        Asistencia.fecha <= fh,
    )
    if id_usuario is not None:
        query = query.filter(Asistencia.id_usuario == id_usuario)

    registros = query.order_by(Asistencia.fecha, Asistencia.id_usuario).limit(limit).all()

    usuarios_ids = list({r.id_usuario for r in registros})
    usuarios_map = {}
    if usuarios_ids:
        usrs = db.query(Usuario).filter(Usuario.id_usuario.in_(usuarios_ids)).all()
        usuarios_map = {u.id_usuario: u.nombre or "" for u in usrs}

    TIPO_LABELS = {
        "TRABAJO": "Trabajo",
        "FESTIVO": "Festivo",
        "VACACION": "Vacación",
        "PERMISO_CON_GOCE": "Permiso c/goce",
        "PERMISO_SIN_GOCE": "Permiso s/goce",
        "INCAPACIDAD": "Incapacidad",
        "FALTA": "Falta",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"
    _encabezado(ws, ["Fecha", "Empleado", "Tipo", "Horas trab.", "Turno completo", "Aplica bono", "Observaciones"])

    for row, r in enumerate(registros, 2):
        tipo_str = getattr(r.tipo, "value", None) or str(r.tipo)
        ws.cell(row=row, column=1, value=r.fecha.strftime("%Y-%m-%d") if r.fecha else "")
        ws.cell(row=row, column=2, value=usuarios_map.get(r.id_usuario, str(r.id_usuario)))
        ws.cell(row=row, column=3, value=TIPO_LABELS.get(tipo_str, tipo_str))
        ws.cell(row=row, column=4, value=round(float(r.horas_trabajadas or 0), 2))
        ws.cell(row=row, column=5, value="Sí" if r.turno_completo else "No")
        ws.cell(row=row, column=6, value="Sí" if r.aplica_bono_puntualidad else "No")
        ws.cell(row=row, column=7, value=(r.observaciones or "")[:500])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"asistencia_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"},
    )


@router.get("/caja")
def exportar_turnos_caja(
    fecha_desde: str | None = Query(None, description="Fecha desde (YYYY-MM-DD)"),
    fecha_hasta: str | None = Query(None, description="Fecha hasta (YYYY-MM-DD)"),
    limit: int = Query(500, ge=1, le=2000),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("ADMIN", "CAJA")),
):
    """Exporta historial de turnos cerrados a Excel con detalle de cortes."""
    query = (
        db.query(CajaTurno)
        .filter(CajaTurno.estado == "CERRADO")
        .options(joinedload(CajaTurno.usuario))
    )
    rol = getattr(current_user.rol, "value", None) or str(current_user.rol)
    if rol == "CAJA":
        query = query.filter(CajaTurno.id_usuario == current_user.id_usuario)
    if fecha_desde:
        fecha_desde_d = datetime.strptime(fecha_desde[:10], "%Y-%m-%d").date()
        query = query.filter(func.date(CajaTurno.fecha_cierre) >= fecha_desde_d)
    if fecha_hasta:
        fecha_hasta_d = datetime.strptime(fecha_hasta[:10], "%Y-%m-%d").date()
        query = query.filter(func.date(CajaTurno.fecha_cierre) <= fecha_hasta_d)

    turnos = query.order_by(CajaTurno.fecha_cierre.desc()).limit(limit).all()

    filas = []
    for t in turnos:
        totales = (
            db.query(Pago.metodo, func.sum(Pago.monto).label("total"))
            .filter(Pago.id_turno == t.id_turno)
            .group_by(Pago.metodo)
            .all()
        )
        efectivo = next((float(tot) for m, tot in totales if m == "EFECTIVO"), 0.0)
        tarjeta = next((float(tot) for m, tot in totales if m == "TARJETA"), 0.0)
        transferencia = next((float(tot) for m, tot in totales if m == "TRANSFERENCIA"), 0.0)
        total_cobros = efectivo + tarjeta + transferencia

        total_gastos = float(
            db.query(func.coalesce(func.sum(GastoOperativo.monto), 0))
            .filter(GastoOperativo.id_turno == t.id_turno)
            .scalar()
            or 0
        )
        total_pagos_prov = float(
            db.query(func.coalesce(func.sum(PagoOrdenCompra.monto), 0))
            .filter(
                PagoOrdenCompra.id_turno == t.id_turno,
                PagoOrdenCompra.metodo == "EFECTIVO",
            )
            .scalar()
            or 0
        )
        efectivo_esperado = (
            float(t.monto_apertura or 0) + efectivo - total_pagos_prov - total_gastos
        )
        monto_cierre = float(t.monto_cierre or 0)
        diferencia = float(t.diferencia) if t.diferencia is not None else None

        usuario_nom = t.usuario.nombre if t.usuario else f"#{t.id_usuario}"
        fecha_cierre_str = t.fecha_cierre.strftime("%Y-%m-%d %H:%M") if t.fecha_cierre else ""
        fecha_apertura_str = t.fecha_apertura.strftime("%Y-%m-%d %H:%M") if t.fecha_apertura else ""

        filas.append({
            "fecha_cierre": fecha_cierre_str,
            "fecha_apertura": fecha_apertura_str,
            "usuario": usuario_nom,
            "apertura": float(t.monto_apertura or 0),
            "efectivo": efectivo,
            "tarjeta": tarjeta,
            "transferencia": transferencia,
            "total_cobros": total_cobros,
            "gastos": total_gastos,
            "pagos_proveedores": total_pagos_prov,
            "efectivo_esperado": efectivo_esperado,
            "monto_contado": monto_cierre,
            "diferencia": diferencia,
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "Turnos de caja"
    _encabezado(ws, [
        "Fecha cierre", "Fecha apertura", "Usuario",
        "Apertura", "Cobros efectivo", "Cobros tarjeta", "Cobros transferencia", "Total cobros",
        "Gastos", "Pagos proveedores", "Efectivo esperado", "Monto contado", "Diferencia",
    ])

    for row, f in enumerate(filas, 2):
        ws.cell(row=row, column=1, value=f["fecha_cierre"])
        ws.cell(row=row, column=2, value=f["fecha_apertura"])
        ws.cell(row=row, column=3, value=f["usuario"])
        ws.cell(row=row, column=4, value=round(f["apertura"], 2))
        ws.cell(row=row, column=5, value=round(f["efectivo"], 2))
        ws.cell(row=row, column=6, value=round(f["tarjeta"], 2))
        ws.cell(row=row, column=7, value=round(f["transferencia"], 2))
        ws.cell(row=row, column=8, value=round(f["total_cobros"], 2))
        ws.cell(row=row, column=9, value=round(f["gastos"], 2))
        ws.cell(row=row, column=10, value=round(f["pagos_proveedores"], 2))
        ws.cell(row=row, column=11, value=round(f["efectivo_esperado"], 2))
        ws.cell(row=row, column=12, value=round(f["monto_contado"], 2))
        ws.cell(row=row, column=13, value=round(f["diferencia"], 2) if f["diferencia"] is not None else "")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"turnos_caja_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"},
    )


@router.get("/sugerencia-compra")
def exportar_sugerencia_compra(
    incluir_cercanos: bool = Query(False, description="Incluir productos cercanos al mínimo"),
    db: Session = Depends(get_db),
    current_user=Depends(require_roles("ADMIN", "CAJA")),
):
    """Exporta la sugerencia de compra (productos con stock bajo) a Excel."""
    from app.models.repuesto import Repuesto

    query = db.query(Repuesto).options(joinedload(Repuesto.proveedor)).filter(
        Repuesto.activo == True,
        Repuesto.eliminado == False,
    )
    if incluir_cercanos:
        query = query.filter(Repuesto.stock_actual <= Repuesto.stock_minimo * 1.2)
    else:
        query = query.filter(Repuesto.stock_actual < Repuesto.stock_minimo)
    repuestos = query.order_by(Repuesto.id_proveedor.asc(), Repuesto.nombre.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Sugerencia compra"
    _encabezado(ws, [
        "Proveedor", "Código", "Nombre", "Stock", "Mín.", "Máx.",
        "Cant. sugerida", "P. compra", "Costo estimado"
    ])

    for row, r in enumerate(repuestos, 2):
        prov = r.proveedor.nombre if r.proveedor else "Sin proveedor"
        cant_min = max(0, r.stock_minimo - r.stock_actual)
        cant_max = max(0, r.stock_maximo - r.stock_actual)
        cant_sug = cant_max if cant_max > 0 else cant_min
        precio = float(r.precio_compra or 0)
        costo = round(precio * cant_sug, 2)
        ws.cell(row=row, column=1, value=prov)
        ws.cell(row=row, column=2, value=r.codigo)
        ws.cell(row=row, column=3, value=r.nombre)
        ws.cell(row=row, column=4, value=r.stock_actual)
        ws.cell(row=row, column=5, value=r.stock_minimo)
        ws.cell(row=row, column=6, value=r.stock_maximo)
        ws.cell(row=row, column=7, value=cant_sug)
        ws.cell(row=row, column=8, value=round(precio, 2))
        ws.cell(row=row, column=9, value=costo)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"sugerencia_compra_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"},
    )
