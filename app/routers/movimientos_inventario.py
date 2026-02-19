"""
Router para Movimientos de Inventario
"""
import csv
import io
import re
import uuid
from pathlib import Path
from decimal import Decimal

from fastapi import APIRouter, Depends, HTTPException, status, Query, File, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from typing import List, Optional
from datetime import datetime
from openpyxl import load_workbook

from app.database import get_db
from app.models.movimiento_inventario import MovimientoInventario, TipoMovimiento
from app.models.repuesto import Repuesto
from app.models.proveedor import Proveedor
from app.models.orden_compra import OrdenCompra
from app.schemas.movimiento_inventario import (
    MovimientoInventarioCreate,
    MovimientoInventarioOut,
    AjusteInventario
)
from app.utils.dependencies import get_current_user
from app.utils.roles import require_roles
from app.models.usuario import Usuario
from app.services.inventario_service import InventarioService
from app.utils.upload import read_file_with_limit

import logging

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/inventario/movimientos",
    tags=["Inventario - Movimientos"]
)

UPLOADS_COMPROBANTES = Path(__file__).resolve().parent.parent.parent / "uploads" / "comprobantes"
ALLOWED_EXT = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".pdf"}
MAX_SIZE_MB = 5

ALLOWED_ENTRADA_MASIVA = {".xlsx", ".csv"}
MAX_ENTRADA_MASIVA_MB = 10
MAX_FILAS_ENTRADA_MASIVA = 500


@router.post("/upload-comprobante")
def subir_comprobante(
    archivo: UploadFile = File(..., description="Imagen o PDF del comprobante de compra"),
    current_user: Usuario = Depends(require_roles("ADMIN", "CAJA", "TECNICO"))
):
    """Sube imagen o PDF de comprobante (factura, recibo). Retorna la URL."""
    UPLOADS_COMPROBANTES.mkdir(parents=True, exist_ok=True)
    ext = Path(archivo.filename or "").suffix.lower()
    if ext not in ALLOWED_EXT:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Formato no permitido. Use: {', '.join(ALLOWED_EXT)}"
        )
    max_bytes = MAX_SIZE_MB * 1024 * 1024
    contenido = read_file_with_limit(archivo.file, max_bytes, MAX_SIZE_MB)
    nombre = f"{uuid.uuid4().hex}{ext}"
    ruta = UPLOADS_COMPROBANTES / nombre
    with open(ruta, "wb") as f:
        f.write(contenido)
    url = f"/uploads/comprobantes/{nombre}"
    return {"url": url}


@router.get("/entrada-masiva/plantilla")
def descargar_plantilla_entrada_masiva(
    current_user: Usuario = Depends(require_roles("ADMIN", "CAJA", "TECNICO"))
):
    """Descarga plantilla Excel para entrada masiva (codigo, cantidad, precio_unitario, referencia, observaciones)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = "Entrada masiva"
    headers = ["codigo", "cantidad", "precio_unitario", "referencia", "observaciones"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
    ws.cell(row=2, column=1, value="MOT-001")
    ws.cell(row=2, column=2, value=10)
    ws.cell(row=2, column=3, value=150.50)
    ws.cell(row=2, column=4, value="FACT-001")
    ws.cell(row=2, column=5, value="Compra proveedor X")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"plantilla_entrada_masiva_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"},
    )


def _parsear_filas_entrada_masiva(contenido: bytes, nombre_archivo: str) -> List[dict]:
    """Extrae filas del archivo Excel o CSV. Retorna lista de dicts con codigo, cantidad, etc."""
    ext = Path(nombre_archivo or "").suffix.lower()
    filas = []
    if ext == ".csv":
        try:
            text = contenido.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = contenido.decode("latin-1")
        reader = csv.DictReader(io.StringIO(text), delimiter=",", skipinitialspace=True)
        for i, row in enumerate(reader):
            row_lower = {k.strip().lower().replace(" ", "_"): v for k, v in (row or {}).items()}
            codigo = str(row_lower.get("codigo", "") or "").strip()
            cantidad_str = str(row_lower.get("cantidad", "") or "").strip()
            if not codigo and not cantidad_str:
                continue
            filas.append({
                "fila": i + 2,
                "codigo": codigo,
                "cantidad": cantidad_str,
                "precio_unitario": str(row_lower.get("precio_unitario", "") or "").strip(),
                "referencia": (str(row_lower.get("referencia", "") or "").strip())[:100],
                "observaciones": (str(row_lower.get("observaciones", "") or "").strip())[:500],
            })
    elif ext == ".xlsx":
        wb = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
        ws = wb.active
        headers = []
        for col in range(1, 10):
            v = ws.cell(row=1, column=col).value
            if v is None:
                break
            headers.append(str(v).strip().lower().replace(" ", "_").replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u"))
        idx_codigo = next((i for i, h in enumerate(headers) if "codigo" in h or h == "cod"), 0)
        idx_cantidad = next((i for i, h in enumerate(headers) if "cantidad" in h), 1)
        idx_precio = next((i for i, h in enumerate(headers) if "precio" in h), 2)
        idx_ref = next((i for i, h in enumerate(headers) if "referencia" in h or "factura" in h), 3)
        idx_obs = next((i for i, h in enumerate(headers) if "observacion" in h or "motivo" in h or "nota" in h), 4)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=max(idx_codigo, idx_cantidad, idx_precio, idx_ref, idx_obs) + 2, values_only=True), start=2):
            row = list(row) if row else []
            codigo = str(row[idx_codigo] or "").strip() if idx_codigo < len(row) else ""
            cantidad_str = str(row[idx_cantidad] or "").strip() if idx_cantidad < len(row) else ""
            if not codigo and not cantidad_str:
                continue
            precio = str(row[idx_precio] or "").strip() if idx_precio < len(row) else ""
            ref = str(row[idx_ref] or "").strip()[:100] if idx_ref < len(row) else ""
            obs = str(row[idx_obs] or "").strip()[:500] if idx_obs < len(row) else ""
            filas.append({"fila": row_idx, "codigo": codigo, "cantidad": cantidad_str, "precio_unitario": precio, "referencia": ref, "observaciones": obs})
        wb.close()
    return filas


@router.post("/entrada-masiva")
def entrada_masiva(
    archivo: UploadFile = File(..., description="Excel o CSV con columnas: codigo, cantidad, precio_unitario, referencia, observaciones"),
    id_proveedor: Optional[int] = Query(None, description="Proveedor por defecto para todas las filas"),
    referencia_global: Optional[str] = Query(None, description="Referencia por defecto (ej: factura)"),
    transaccional: bool = Query(False, description="Si True, falla todo ante el primer error (todo o nada)"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("ADMIN", "CAJA", "TECNICO"))
):
    """
    Registra entradas masivas desde Excel o CSV.
    Columnas esperadas: codigo, cantidad, precio_unitario (opc), referencia (opc), observaciones (opc).
    Máximo 500 filas. Si transaccional=True, falla todo ante el primer error (todo o nada).
    """
    ext = Path(archivo.filename or "").suffix.lower()
    if ext not in ALLOWED_ENTRADA_MASIVA:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Formato no permitido. Use: {', '.join(ALLOWED_ENTRADA_MASIVA)}"
        )
    max_bytes = MAX_ENTRADA_MASIVA_MB * 1024 * 1024
    contenido = read_file_with_limit(archivo.file, max_bytes, MAX_ENTRADA_MASIVA_MB)
    try:
        filas = _parsear_filas_entrada_masiva(contenido, archivo.filename or "")
    except Exception as e:
        logger.warning(f"Error parseando archivo entrada masiva: {e}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"No se pudo leer el archivo: {str(e)}"
        )
    if not filas:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El archivo no contiene filas válidas. Verifica que tenga encabezados: codigo, cantidad"
        )
    if len(filas) > MAX_FILAS_ENTRADA_MASIVA:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"El archivo tiene {len(filas)} filas. Máximo permitido: {MAX_FILAS_ENTRADA_MASIVA}. "
                   f"Divida el archivo o procese en lotes.",
        )
    if id_proveedor:
        prov = db.query(Proveedor).filter(
            Proveedor.id_proveedor == id_proveedor,
            Proveedor.activo == True
        ).first()
        if not prov:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Proveedor no encontrado o inactivo. Verifique el ID del proveedor."
            )
    def _fallar_si_transaccional(msg: str, fila: int, codigo: str):
        if transaccional:
            db.rollback()
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Error en fila {fila} (código {codigo or '(vacío)'}): {msg}"
            )

    procesados = 0
    errores = []
    for item in filas:
        codigo = (item.get("codigo") or "").strip()
        cantidad_str = (item.get("cantidad") or "").strip()
        fila_num = item.get("fila", 0)
        if not codigo:
            _fallar_si_transaccional("Código vacío", fila_num, codigo)
            errores.append({"fila": fila_num, "codigo": codigo or "(vacío)", "error": "Código vacío"})
            continue
        if not cantidad_str:
            _fallar_si_transaccional("Cantidad vacía", fila_num, codigo)
            errores.append({"fila": fila_num, "codigo": codigo, "error": "Cantidad vacía"})
            continue
        try:
            cantidad = float(cantidad_str)
        except (ValueError, TypeError):
            _fallar_si_transaccional(f"Cantidad inválida: {cantidad_str}", fila_num, codigo)
            errores.append({"fila": fila_num, "codigo": codigo, "error": f"Cantidad inválida: {cantidad_str}"})
            continue
        if cantidad < 0.001:
            _fallar_si_transaccional("Cantidad debe ser al menos 0.001", fila_num, codigo)
            errores.append({"fila": fila_num, "codigo": codigo, "error": "Cantidad debe ser al menos 0.001"})
            continue
        repuesto = db.query(Repuesto).filter(
            Repuesto.codigo.ilike(codigo),
            Repuesto.eliminado == False
        ).first()
        if not repuesto:
            _fallar_si_transaccional("Repuesto no encontrado", fila_num, codigo)
            errores.append({"fila": fila_num, "codigo": codigo, "error": "Repuesto no encontrado"})
            continue
        if not repuesto.activo:
            _fallar_si_transaccional("Repuesto inactivo", fila_num, codigo)
            errores.append({"fila": fila_num, "codigo": codigo, "error": "Repuesto inactivo"})
            continue
        precio_val = None
        if item.get("precio_unitario"):
            try:
                precio_val = Decimal(str(item["precio_unitario"]).replace(",", "."))
                if precio_val < 0:
                    precio_val = None
            except (ValueError, Exception):
                pass
        ref = item.get("referencia") or referencia_global
        motivo = item.get("observaciones")
        mov_data = {
            "id_repuesto": repuesto.id_repuesto,
            "tipo_movimiento": TipoMovimiento.ENTRADA,
            "cantidad": cantidad,
            "precio_unitario": precio_val or repuesto.precio_compra,
            "referencia": ref or None,
            "motivo": motivo or None,
            "id_proveedor": id_proveedor,
            "fecha_adquisicion": datetime.utcnow().date(),
        }
        try:
            movimiento = MovimientoInventarioCreate(**mov_data)
            InventarioService.registrar_movimiento(
                db=db,
                movimiento=movimiento,
                id_usuario=current_user.id_usuario,
                autocommit=not transaccional,
            )
            procesados += 1
        except Exception as e:
            if transaccional:
                db.rollback()
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Error en fila {item.get('fila', 0)} (código {codigo}): {str(e)[:200]}"
                )
            errores.append({"fila": item.get("fila", 0), "codigo": codigo, "error": str(e)[:200]})
    if transaccional and procesados > 0:
        db.commit()
    return {
        "mensaje": f"Procesadas {procesados} entradas",
        "procesados": procesados,
        "total_filas": len(filas),
        "errores": errores,
    }


@router.post("/", response_model=MovimientoInventarioOut, status_code=status.HTTP_201_CREATED)
def registrar_movimiento(
    movimiento: MovimientoInventarioCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("ADMIN", "CAJA", "TECNICO"))
):
    """
    Registra un movimiento de inventario (entrada, salida, merma, etc.).
    
    Requiere rol: ADMIN, CAJA o TECNICO
    
    Tipos de movimiento:
    - ENTRADA: Compra o devolución
    - SALIDA: Venta o uso en servicio
    - AJUSTE+: Corrección al alza
    - AJUSTE-: Corrección a la baja
    - MERMA: Pérdida o daño
    """
    try:
        nuevo_movimiento = InventarioService.registrar_movimiento(
            db=db,
            movimiento=movimiento,
            id_usuario=current_user.id_usuario
        )
        
        return nuevo_movimiento
    
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        logger.error(f"Error al registrar movimiento: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al registrar el movimiento de inventario"
        )


@router.post("/ajuste", response_model=MovimientoInventarioOut, status_code=status.HTTP_201_CREATED)
def ajustar_inventario(
    ajuste: AjusteInventario,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("ADMIN", "CAJA"))
):
    """
    Ajusta el inventario a un valor específico.
    
    Requiere rol: ADMIN o CAJA
    
    Útil para correcciones de inventario físico.
    """
    try:
        movimiento_ajuste = InventarioService.ajustar_inventario(
            db=db,
            ajuste=ajuste,
            id_usuario=current_user.id_usuario
        )
        
        return movimiento_ajuste
    
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        logger.error(f"Error al ajustar inventario: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al ajustar el inventario"
        )


@router.get("/")
def listar_movimientos(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    id_repuesto: Optional[int] = Query(None, description="Filtrar por repuesto"),
    tipo_movimiento: Optional[TipoMovimiento] = Query(None, description="Filtrar por tipo"),
    fecha_desde: Optional[datetime] = Query(None, description="Fecha desde (YYYY-MM-DD)"),
    fecha_hasta: Optional[datetime] = Query(None, description="Fecha hasta (YYYY-MM-DD)"),
    id_usuario: Optional[int] = Query(None, description="Filtrar por usuario"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Lista todos los movimientos de inventario con filtros opcionales.
    
    Filtros disponibles:
    - id_repuesto: Movimientos de un repuesto específico
    - tipo_movimiento: ENTRADA, SALIDA, AJUSTE+, AJUSTE-, MERMA
    - fecha_desde: Movimientos desde esta fecha
    - fecha_hasta: Movimientos hasta esta fecha
    - id_usuario: Movimientos realizados por un usuario
    """
    query = db.query(MovimientoInventario)
    
    # Aplicar filtros
    if id_repuesto:
        query = query.filter(MovimientoInventario.id_repuesto == id_repuesto)
    
    if tipo_movimiento:
        query = query.filter(MovimientoInventario.tipo_movimiento == tipo_movimiento)
    
    if fecha_desde:
        query = query.filter(MovimientoInventario.fecha_movimiento >= fecha_desde)
    
    if fecha_hasta:
        query = query.filter(MovimientoInventario.fecha_movimiento <= fecha_hasta)
    
    if id_usuario:
        query = query.filter(MovimientoInventario.id_usuario == id_usuario)
    
    # Ordenar por fecha descendente y cargar relaciones para respuesta
    total = query.count()
    movimientos = query.options(
        joinedload(MovimientoInventario.repuesto),
        joinedload(MovimientoInventario.usuario),
    ).order_by(
        MovimientoInventario.fecha_movimiento.desc()
    ).offset(skip).limit(limit).all()

    return {
        "movimientos": movimientos,
        "total": total,
        "pagina": skip // limit + 1 if limit > 0 else 1,
        "total_paginas": (total + limit - 1) // limit if limit > 0 else 1,
        "limit": limit,
    }


@router.get("/repuesto/{id_repuesto}", response_model=List[MovimientoInventarioOut])
def historial_repuesto(
    id_repuesto: int,
    limite: int = Query(50, le=200, description="Número máximo de registros"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Obtiene el historial completo de movimientos de un repuesto específico.
    """
    # Verificar que el repuesto existe
    repuesto = db.query(Repuesto).filter(
        Repuesto.id_repuesto == id_repuesto
    ).first()
    
    if not repuesto:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Repuesto con ID {id_repuesto} no encontrado"
        )
    
    movimientos = db.query(MovimientoInventario).filter(
        MovimientoInventario.id_repuesto == id_repuesto
    ).order_by(
        MovimientoInventario.fecha_movimiento.desc()
    ).limit(limite).all()
    
    # Enriquecer referencias OC-{id} con numero (OC-YYYYMMDD-NNNN) y comprobante
    ids_oc = set()
    for m in movimientos:
        if m.referencia and re.match(r"^OC-(\d+)$", (m.referencia or "").strip()):
            ids_oc.add(int(re.match(r"^OC-(\d+)$", m.referencia.strip()).group(1)))
    mapa_oc = {}
    if ids_oc:
        for oc in db.query(OrdenCompra).filter(
            OrdenCompra.id_orden_compra.in_(ids_oc)
        ).all():
            mapa_oc[oc.id_orden_compra] = {
                "numero": oc.numero or f"OC-{oc.id_orden_compra}",
                "comprobante_url": (oc.comprobante_url or "").strip() or None,
            }
    for m in movimientos:
        if m.referencia:
            match = re.match(r"^OC-(\d+)$", (m.referencia or "").strip())
            if match:
                id_oc = int(match.group(1))
                if id_oc in mapa_oc:
                    m.referencia = mapa_oc[id_oc]["numero"]
                    if not (m.imagen_comprobante_url or "").strip() and mapa_oc[id_oc]["comprobante_url"]:
                        m.imagen_comprobante_url = mapa_oc[id_oc]["comprobante_url"]
    
    return movimientos


@router.get("/{id_movimiento}", response_model=MovimientoInventarioOut)
def obtener_movimiento(
    id_movimiento: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Obtiene los detalles de un movimiento específico.
    """
    movimiento = db.query(MovimientoInventario).filter(
        MovimientoInventario.id_movimiento == id_movimiento
    ).first()
    
    if not movimiento:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Movimiento con ID {id_movimiento} no encontrado"
        )
    
    return movimiento


@router.get("/estadisticas/resumen")
def obtener_estadisticas_movimientos(
    fecha_desde: Optional[datetime] = Query(None),
    fecha_hasta: Optional[datetime] = Query(None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_roles("ADMIN", "CAJA"))
):
    """
    Obtiene estadísticas de movimientos de inventario.
    
    Requiere rol: ADMIN o CAJA
    """
    from sqlalchemy import func
    
    query = db.query(
        MovimientoInventario.tipo_movimiento,
        func.count(MovimientoInventario.id_movimiento).label("total_movimientos"),
        func.sum(MovimientoInventario.cantidad).label("total_cantidad"),
        func.sum(MovimientoInventario.costo_total).label("total_costo")
    )
    
    if fecha_desde:
        query = query.filter(MovimientoInventario.fecha_movimiento >= fecha_desde)
    
    if fecha_hasta:
        query = query.filter(MovimientoInventario.fecha_movimiento <= fecha_hasta)
    
    estadisticas = query.group_by(
        MovimientoInventario.tipo_movimiento
    ).all()
    
    resultado = {
        "periodo": {
            "desde": fecha_desde.isoformat() if fecha_desde else None,
            "hasta": fecha_hasta.isoformat() if fecha_hasta else None
        },
        "por_tipo": [
            {
                "tipo": stat.tipo_movimiento,
                "total_movimientos": stat.total_movimientos,
                "total_cantidad": stat.total_cantidad,
                "total_costo": float(stat.total_costo) if stat.total_costo else 0
            }
            for stat in estadisticas
        ]
    }
    
    return resultado
